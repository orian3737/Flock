import csv
from datetime import datetime
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import AnimalClass, FeedType, FeedingEvent, Flock, InventoryTransaction, ProductionLog, Revenue, User
from app.services.financial_service import get_farm_summary, get_user_flock_pl


def generate_csv(user_id, report_type, flock_ids, start_date, end_date):
    output = StringIO()
    writer = csv.writer(output)
    rows = preview_rows(user_id, report_type, flock_ids, start_date, end_date, limit=None)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return output


def generate_pdf(user_id, report_type, flock_ids, start_date, end_date):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    user = User.query.get(user_id)

    story.append(Paragraph(user.farm_name if user else "Flock", styles["Title"]))
    story.append(Paragraph(f"{report_type.replace('_', ' ').title()} report", styles["Heading2"]))
    story.append(Paragraph(f"{start_date.isoformat()} to {end_date.isoformat()}", styles["Normal"]))
    story.append(Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 20))

    sections = _report_sections(report_type)
    for index, section in enumerate(sections):
        rows = preview_rows(user_id, section, flock_ids, start_date, end_date, limit=None)
        story.append(Paragraph(section.replace("_", " ").title(), styles["Heading2"]))
        if rows:
            table = Table(rows, repeatRows=1)
            table.setStyle(_pdf_table_style())
            story.append(table)
        else:
            story.append(Paragraph("No rows found.", styles["Normal"]))
        if index < len(sections) - 1:
            story.append(PageBreak())

    doc.build(story, onFirstPage=_page_number, onLaterPages=_page_number)
    output.seek(0)
    return output


def generate_xlsx(user_id, flock_ids, start_date, end_date):
    output = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = {
        "Summary": _summary_rows(user_id, start_date, end_date),
        "Feeding Log": preview_rows(user_id, "feeding_log", flock_ids, start_date, end_date, limit=None),
        "Production": preview_rows(user_id, "production_log", flock_ids, start_date, end_date, limit=None),
        "Inventory": preview_rows(user_id, "inventory", flock_ids, start_date, end_date, limit=None),
        "Financials": preview_rows(user_id, "financial_summary", flock_ids, start_date, end_date, limit=None),
    }

    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
        _style_worksheet(worksheet)

    workbook.save(output)
    output.seek(0)
    return output


def preview_rows(user_id, report_type, flock_ids=None, start_date=None, end_date=None, limit=10):
    section = report_type if report_type != "full" else "feeding_log"
    if section == "feeding_log":
        rows = _feeding_rows(user_id, flock_ids, start_date, end_date)
    elif section == "production_log":
        rows = _production_rows(user_id, flock_ids, start_date, end_date)
    elif section == "financial_summary":
        rows = _financial_rows(user_id, start_date, end_date)
    elif section == "inventory":
        rows = _inventory_rows(user_id)
    else:
        rows = _feeding_rows(user_id, flock_ids, start_date, end_date)
    return rows[: limit + 1] if limit is not None else rows


def _feeding_rows(user_id, flock_ids, start_date, end_date):
    query = (
        FeedingEvent.query.join(Flock)
        .join(Flock.breed)
        .join(AnimalClass)
        .filter(AnimalClass.user_id == user_id, FeedingEvent.date >= start_date, FeedingEvent.date <= end_date)
        .order_by(FeedingEvent.date, FeedingEvent.timestamp)
    )
    if flock_ids:
        query = query.filter(FeedingEvent.flock_id.in_(flock_ids))

    rows = [["Date", "Time", "Flock", "Breed", "Feed", "Weight", "Wt/Bird", "Cost", "Cost/Bird", "Method"]]
    for event in query.all():
        rows.append(
            [
                event.date.isoformat(),
                event.timestamp.strftime("%H:%M") if event.timestamp else "",
                event.flock.name,
                event.flock.breed.name,
                event.feed_type.name,
                round(event.total_weight, 2),
                round(event.weight_per_bird, 3),
                round(event.cost_total, 2),
                round(event.cost_per_bird, 3),
                event.input_method,
            ]
        )
    return rows


def _production_rows(user_id, flock_ids, start_date, end_date):
    query = (
        ProductionLog.query.join(Flock)
        .join(Flock.breed)
        .join(AnimalClass)
        .filter(AnimalClass.user_id == user_id, ProductionLog.date >= start_date, ProductionLog.date <= end_date)
        .order_by(ProductionLog.date)
    )
    if flock_ids:
        query = query.filter(ProductionLog.flock_id.in_(flock_ids))
    rows = [["Date", "Flock", "Egg Count", "Water Consumed", "Avg Weight", "Notes"]]
    for log in query.all():
        rows.append([log.date.isoformat(), log.flock.name, log.egg_count, log.water_consumed, log.avg_weight, log.notes])
    return rows


def _financial_rows(user_id, start_date, end_date):
    rows = [["Flock", "Designation", "Headcount", "Feed Cost", "Revenue", "Net P&L", "Cost/Bird", "Cost/Dozen"]]
    for flock in get_user_flock_pl(user_id, start_date, end_date):
        rows.append(
            [
                flock["name"],
                flock["designation"],
                flock["headcount"],
                flock["total_feed_cost"],
                flock["total_revenue"],
                flock["net_pl"],
                flock["cost_per_bird"],
                flock["cost_per_dozen"],
            ]
        )
    return rows


def _inventory_rows(user_id):
    rows = [["Feed", "Unit", "On Hand", "Par Level", "Bag Weight", "Bag Price", "Cost/Lb"]]
    for feed in FeedType.query.filter_by(user_id=user_id).order_by(FeedType.name).all():
        rows.append(
            [
                feed.name,
                feed.unit,
                round(feed.current_on_hand, 2),
                round(feed.par_level, 2),
                round(feed.bag_weight, 2),
                round(feed.bag_price, 2),
                round(feed.cost_per_lb, 4),
            ]
        )
    return rows


def _summary_rows(user_id, start_date, end_date):
    summary = get_farm_summary(user_id, start_date, end_date)
    return [
        ["Metric", "Value"],
        ["Total Feed Cost", summary["total_feed_cost"]],
        ["Total Revenue", summary["total_revenue"]],
        ["Net P&L", summary["net_pl"]],
        ["Top Cost Flock", summary["top_cost_flock"]["name"] or ""],
    ]


def _report_sections(report_type):
    if report_type == "full":
        return ["feeding_log", "production_log", "financial_summary", "inventory"]
    return [report_type]


def _pdf_table_style():
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A1A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7E8D7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F1F8F1"), colors.white]),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]
    )


def _page_number(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(560, 20, f"Page {doc.page}")
    canvas.restoreState()


def _style_worksheet(worksheet):
    header_fill = PatternFill("solid", fgColor="1A3A1A")
    alt_fill = PatternFill("solid", fgColor="F1F8F1")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=12)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        if row_index % 2 == 0:
            for cell in row:
                cell.fill = alt_fill
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00" if any(word in str(worksheet.cell(1, cell.column).value) for word in ["Cost", "Revenue", "P&L"]) else "0.00"
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 32)
